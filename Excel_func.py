import openpyxl, glob, os, json,datetime
from openpyxl import Workbook as wb
from win32com import client
import Function as fn
from docx import Document


def test():
	projekt_ID = '22/41'
	projekt = fn.projekt(projekt_ID)
	info = projekt.hledani_info_mesta()
	
	excel = excel_manipulation(info)
	# stitky = excel.file_find_create('Titulní strany')
	stitky = excel.stitky_pdf()



class excel_manipulation:
	def __init__(self,info_dict):
		self.project_dict = info_dict
		self.user_name = os.getlogin()
		cwd = os.getcwd()

		#load potrebne odkazy	
		self.Titulni_strany = rf'{cwd}\Dokumenty\Titulní strany a příprava R.xlsx'
		self.project_f_dir = self.project_dict["project_f_dir"]
		self.seznam_doc_directory = rf'{self.project_f_dir}\dokumentace na SÚ\dokladová část'
		

		#open setting na odkzay
		cwd = os.path.dirname(os.path.realpath(__file__))
		with open(rf"{cwd}\setting.json","r",encoding = 'utf-8') as f:
			file = f.read()
			self.setting = json.loads(file)
		
	def open_main_xl(self): #open zadani list na vyplneni info -> return aktivni excel list
		wb = openpyxl.load_workbook(self.Titulni_strany) #load seznam projectu file
		ws = wb.sheetnames

		for index,sheet in enumerate(ws): 
			if self.setting["Titulni_strany"]["zadani_list"] == sheet:
				company = 'R-built'
				wb.active = index
				ws= wb.active
				self.ws = ws
				self.wb = wb

		return	

	def vypln_zad_list(self): #vyplneno ZAD list v excel
		project_f_dir = self.project_dict["project_f_dir"]
		pruvodni_zprava_directory = rf'{project_f_dir}\dokumentace na SÚ\textová část'
		doc_SU_path = rf'{project_f_dir}\dokumentace na SÚ'
		
		datum_ = self.datum_vydani()


		self.ws[str(self.setting["Titulni_strany"]["Projektant"])] = self.project_dict["jmeno_projektanta"]
		self.ws[str(self.setting["Titulni_strany"]["Nazev_stavby"])] = self.project_dict["nazev_projektu"]
		self.ws[str(self.setting["Titulni_strany"]["interni_cislo"])] = f'{self.project_dict["cislo_projektu"]}/{self.project_dict["rok_projektu"]}'
		self.ws[str(self.setting["Titulni_strany"]["datum"])] = datum_
		self.ws[str(self.setting["Titulni_strany"]["Cislo_stavby"])] = self.project_dict["kod_projektu"]
		self.ws[str(self.setting["Titulni_strany"]["Tel_cislo_projektant"])] = self.project_dict["Tel_cislo"]

		out_excel_file = rf'{doc_SU_path}\Titulní strany a příprava R.xlsx'	
		self.wb.save(filename=out_excel_file)
		self.wb.close()

		return out_excel_file 

	def stitky_pdf (self): 
		#hledani existujici titulni stranky
		list_tit_str_path = self.file_find_create('Titulní strany')
		project_f_dir = self.project_dict["project_f_dir"]
		doc_SU = project_f_dir + r"\dokumentace na SÚ"

		datum_ = self.datum_vydani()

		#jestli existuje
		if [os.path.exists(path) for path in list_tit_str_path]:
			
			#existujici titulni stranka
			for excel_file in list_tit_str_path:
				#otevreni titulni strany
				work_b = openpyxl.load_workbook(excel_file) #load seznam projectu file
				work_s = work_b.sheetnames		
				
				excel = client.Dispatch("Excel.Application")
				excel.Visible = False
								
				for index,sheet in enumerate(work_s): 
					if self.setting["Titulni_strany"]["stitky"] == sheet:
						exist = True
						work_b.close()
						
						#otevrit existujici
						sheets = excel.Workbooks.Open(excel_file)
						work_sheets = sheets.Worksheets[index]
						work_sheets.ExportAsFixedFormat(0, rf'{doc_SU}\ŠTÍTKY.pdf')
						sheets.Close(True)

						return 'Hotovo'

				work_b.close()

			#kdyz budes stara verze titulni strany zvlast export stitky
			if 'exist' not in locals():
				cwd = os.path.dirname(os.path.realpath(__file__))
				stitky = rf"{cwd}\Dokumenty\ŠTÍTKY.xlsx"
				work_b = openpyxl.load_workbook(stitky)
				work_s = work_b.active

				excel = client.Dispatch("Excel.Application")
				excel.Visible = False

				work_s[str(self.setting["stitky"]["Projektant"])] = self.project_dict["jmeno_projektanta"]
				work_s[str(self.setting["stitky"]["Nazev_stavby"])] = self.project_dict["nazev_projektu"]
				work_s[str(self.setting["stitky"]["interni_cislo"])] = f'{self.project_dict["cislo_projektu"]}/{self.project_dict["rok_projektu"]}'
				work_s[str(self.setting["stitky"]["datum"])] = datum_
				work_s[str(self.setting["stitky"]["Cislo_stavby"])] = self.project_dict["kod_projektu"]
				
				out_excel_file = rf'C:\Users\{self.user_name}\Desktop\ŠTÍTKY.xlsx'	
				work_b.save(filename=out_excel_file)
				work_b.close()

				#pdf format export
				sheets = excel.Workbooks.Open(out_excel_file)
				work_sheets = sheets.Worksheets[0]
				work_sheets.ExportAsFixedFormat(0, rf'{doc_SU}\ŠTÍTKY.pdf')
				sheets.Close(True)

				os.remove(out_excel_file)

		else:
			raise Exception("Nenalezena titulní strany")

		return 'Done'


	def doklad1 (self):

		#open vzorovy excel
		self.open_main_xl()
		#vyplnit vzorovy excel
		self.vyplneny_excel_file = self.vypln_zad_list()

		seznam_doc_directory = self.seznam_doc_directory

		def read_seznam_doc (seznam_doc_directory):

			seznam_dict = {
				"file" : "Seznam dokumentů",
				"distribuce" : "ČEZ Distribuce a.s.",
				"ICT" : "ČEZ ICT Services. a.s.",
				"TPS" : "Telco Pro Services a.s.",
				"cetin" : "CETIN",
				"gasnet" : "GasNet, s.r.o. ",
				"CRA" : "České Radiokomunikace a.s.",
				"tmobile" : "T-mobile a.s.",
				"vodafone" : "Vodafone a.s.",
				"scvk" : "SčVK a.s.",
				"cd" : "ČD telematika"
			}

			info_seznam = {}

			#find seznam dokumetu
			

			os.chdir(seznam_doc_directory)

			for excel_file in glob.glob('*.xlsx'):
				if seznam_dict["file"] in excel_file:
					seznam_doc_directory = rf"{seznam_doc_directory}\{excel_file}"

			wb = openpyxl.load_workbook(seznam_doc_directory, read_only=True) #load seznam projectu file
			ws = wb.active

			
			#hledani bunek a informaci v tabulkach
			for row in ws.rows:	
				for cell in row:
					for subject in seznam_dict.values():

						sub_dict = {}
						if str(cell.value) == subject:
							
							podmin = {
								"ano" : "střet",
								"ne" : "nenachází"
							}
							try:
								platnost = ws.cell(row=cell.row, column=4).value.strftime("%d.%m.%Y")
								if platnost is not None:
									sub_dict["platnost"] = platnost
							except:
								sub_dict["platnost"] = ""
							
								

							podminka = ws.cell(row=cell.row, column=5).value

							sub_dict["podminka"] = podmin[f"{podminka}"]

							znacka = ws.cell(row=cell.row, column=6).value
							sub_dict["znacka"] = znacka

							info_seznam[f"{subject}"] = sub_dict
							
			return info_seznam

		seznam_info = read_seznam_doc(seznam_doc_directory)

		#najit titulni stranu
		list_tit_str_path = self.file_find_create('Titulní strany')
		project_f_dir = self.project_dict["project_f_dir"]
		doc_SU = project_f_dir + r"\dokumentace na SÚ"

		if [os.path.exists(path) for path in list_tit_str_path]:
			
			for excel in list_tit_str_path:
				
				wb = openpyxl.load_workbook(excel) #load seznam projectu file
				ws = wb.sheetnames
				exist = False
				for index,sheet in enumerate(ws):

					if "SoBS tabulka" == sheet:
						sobs = True
						wb.active = index
						ws = wb.active
					
						ws["B3"] = self.project_dict["k.u."]

					if "H1" == sheet:
						H1 = True
						exist = True
						company = 'R-built'
						wb.active = index
						ws= wb.active
						
						for index, tupl in enumerate(seznam_info.items()):
							subjekt, info_dict = tupl
							# print(info_dict)
							ws[f"B{7+index}"] = subjekt
							ws[f"C{7+index}"] = info_dict["platnost"]
							ws[f"D{7+index}"] = info_dict["znacka"]
							ws[f"E{7+index}"] = info_dict["podminka"]


				if 'sobs' and 'H1' in locals():
					
					out_excel_file = excel
					wb.save(filename=out_excel_file)
					wb.close()

					return 'Hotovo'

				wb.close()				

			if exist == False:
				raise Exception("Nenalezena Titulní strany")
		
		return 'Path not exist'


	def file_find_create(self, podminka):
		os.chdir(self.project_dict["project_f_dir"])
		list_file = []
		for root, dirs, files in os.walk(".",topdown=True):
			
			for name in files:
				if podminka in name:
					exist = True
					titul_stranka_path = os.path.join(root, name)			
					titul_stranka_path = os.path.abspath(titul_stranka_path)
					list_file.append(titul_stranka_path)
		

		if 'exist' not in locals():
			titul_stranka_path = self.project_dict["project_f_dir"] + r"\dokumentace na SÚ\Titulní strany a příprava R.xlsx"
			print("vytvari se...")
		
		return list_file

	def datum_vydani(self):
		project_f_dir = self.project_dict["project_f_dir"]
		pruvodni_zprava_directory = rf'{project_f_dir}\dokumentace na SÚ\textová část'

		doc_files = glob.glob(rf'{pruvodni_zprava_directory}\*.docx')
		
		for file_name in doc_files:
			# print (file_name)
			if 'A_Průvodní' or 'Technicka' in file_name:
				pruvodni_zprava_directory = file_name
				# print ("Hotovo: ", pruvodni_zprava_directory)
			else :
				print ("Pruvodni zprava nenalezena")
				return None

		with open(pruvodni_zprava_directory,'rb') as f:
				document = Document(f)

		for index,paragraph in enumerate(document.paragraphs):
			if 'Datum vydání' in paragraph.text:
				datum_vydani = paragraph.text.replace("Datum vydání", "").replace(":", "").lstrip()
				
		return	datum_vydani
if __name__ == '__main__':
	test()